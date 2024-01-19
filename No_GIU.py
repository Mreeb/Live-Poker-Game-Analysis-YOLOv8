# Install These Libraries in the Python Terminal
# pip install pyautogui webbrowser tkinter ultralytics openpyxl

import pyautogui
import time
import webbrowser
import tkinter as tk
from tkinter import Button
from ultralytics import YOLO
import openpyxl
import os
from openpyxl.styles import PatternFill, Border, Side

url = "https://roobet.com/game?src=aHR0cHM6Ly90ZWtob3U1LWRrMi5wcGdhbWVzLm5ldC9nczJjL3BsYXlHYW1lLmRvP2tleT10b2tlbiUzRDQ5ODdhNTIxLWJkODctNDhkOC04YTM3LWQ2OGY5YTg0OTI2OCUyNTNBMzU4ZjIyMTgxNTAwZjVmODRiNjU3MzA1NDhmNjE2ZjU1ZmIzODU3ZiUyNnN5bWJvbCUzRDUyMCUyNnRlY2hub2xvZ3klM0RINSUyNnBsYXRmb3JtJTNETU9CSUxFJTI2bGFuZ3VhZ2UlM0RlbiUyNmNhc2hpZXJVcmwlM0RodHRwcyUzQSUyRiUyRnJvb2JldC5jb20lMkYlM0Ztb2RhbCUzRGNhc2hpZXIlMjZ0YWIlM0RkZXBvc2l0JTI2bG9iYnlVcmwlM0RodHRwcyUzQSUyRiUyRnJvb2JldC5jb20mc3R5bGVuYW1lPXRla2hfcm9vYmV0JmNvdW50cnk9Q0EmaXNHYW1lVXJsQXBpQ2FsbGVkPXRydWU"
#url = "https://www.google.com"

# chrome_path = "C:/Program Files/Google/Chrome/Application/chrome.exe"
chrome_path = "/snap/bin/opera"

url_to_open = url
output_folder = "frames_folder"

webbrowser.register('chrome', None, webbrowser.BackgroundBrowser(chrome_path))
webbrowser.get('chrome').open(url_to_open)

time.sleep(10)
def detection_func(ss):
    model = YOLO('Detector.pt')
    resultss = model(ss, imgsz=512)
    res = []

    for results in resultss:
        try:
            if results.boxes.cls.numel() == 0:
                res.append("000000")
            else:
                val = int(results.boxes.cls)
                if val == 4 or val == 3:
                    res.append("FF0000")
                elif val == 2 or val == 1:
                    res.append("00FF00")
                elif val == 0:
                    res.append("FFFF00")

        except (IndexError, AttributeError, TypeError, ValueError):
            res.append("FF0000")

    return res

def create_and_fill_colors(user_colors, sheet_name="Sheet1"):
    file_name = 'New_Results2.xlsx'

    if os.path.isfile(file_name):
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook[sheet_name]
        existing_rows = worksheet.max_row
        row_num = existing_rows + 1
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name
        row_num = 2

        for col_num, user_header in enumerate(range(1, len(user_colors) + 1), start=1):
            worksheet.cell(row=1, column=col_num, value=f'user{user_header}')

    for col_num, color in enumerate(user_colors, start=1):
        if isinstance(color, str):
            color_rgb = color
        else:
            color_rgb = ''.join(format(c, '02X') for c in color)
        fill = PatternFill(start_color=color_rgb, end_color=color_rgb, fill_type='solid')

        cell = worksheet.cell(row=row_num, column=col_num)
        cell.fill = fill

        border = Border(left=Side(style='thin', color='FFFFFFFF'),
                        right=Side(style='thin', color='FFFFFFFF'),
                        top=Side(style='thin', color='FFFFFFFF'),
                        bottom=Side(style='thin', color='FFFFFFFF'))
        cell.border = border

    workbook.save(file_name)

def capture_and_save_Excel(region):
    screenshot = pyautogui.screenshot(region=region)
    width, height = screenshot.size
    section_width = width // 7
    sub_images = []

    for i in range(7):
        left = i * section_width
        upper = 0
        right = left + section_width
        lower = height
        section = screenshot.crop((left, upper, right, lower))
        sub_images.append(section)

    excel_values = detection_func(sub_images)
    return excel_values

if __name__ == "__main__":
    save_path = "captured_frames/"
    try:
        os.makedirs(save_path)
    except FileExistsError:
        pass

    region = (300, 540, 845, 225)

    old_values = []
    while True:
        excel_values = capture_and_save_Excel(region=region)

        if old_values != excel_values:
            create_and_fill_colors(excel_values)

        old_values = excel_values
